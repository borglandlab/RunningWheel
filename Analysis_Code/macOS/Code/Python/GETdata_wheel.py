
import os, re, GETdata_wheel
from openpyxl import load_workbook, Workbook
from inbox import get_inbox
from send_wheelalert import send_wheelalert
from datetime import datetime, timedelta
import pytz

#This will need to be as long as the number of spinners you have. This variable is used in Wheel_RunFirst and GetData_wheel functions.
Spinner_List = ['Spinner_1',
'Spinner_2',
'Spinner_3',
'Spinner_4',
'Spinner_5',
'Spinner_6',
'Spinner_7',
'Spinner_8']

Spinner_Email = 'email address'	#This is the email where your data from the running wheels is being sent

SavedData_Directory = '/Users/<USERNAME>/Desktop/RunningWheel/Wheel_Data'	#This will need to be the folder directory where the data is saved.
Code_Directory = '/Users/<USERNAME>/Desktop/RunningWheel/Python_Code'	#This will need to be the folder directory where the code is kept.

EmailAlert_hours = 3 	#will send an alert if this number of hours passes without recieving an email

def Wheel_RunFirst():
	#This function sets everything up. It only has to be run once at the very beginning before the first download.
	# This will create my excel files that will contain:
	# 1. A directory in an excel file with a tab for each wheel
	# 2. An excel file for each wheel to store data

	os.chdir(SavedData_Directory)	#This is the directory where you want your data to be stored after download

	wb_directory = Workbook()	#Creats a new excel workbook that will be used as a directory for all spinners

	ws_length = wb_directory.create_sheet('Length', 0)	#New sheet called "Length"
	ws_length.cell(row=1, column=1).value = 'Spinner'	#Column titles "Spinner", "Files", and "Length"
	ws_length.cell(row=1, column=2).value = 'Files'
	ws_length.cell(row=1, column=3).value = 'Length'

	for n in range(len(Spinner_List)):
		ws_length.cell(row=n+2, column=1).value = Spinner_List[n]	#The name of the spinner (ie. Spinner_1... Spinner_8)
		ws_directory = wb_directory.create_sheet(Spinner_List[n], n+1)	#Creates a sheet that will be th directory for this specific spinner
		ws_directory.cell(row=1, column=1).value = 'File'	#Which the data is on (ie. '_1.xlsx', '_2.xlsx', '_3.xlsx', ect.)
		ws_directory.cell(row=1, column=2).value = 'Tab_Name'	#Tab name will be the date and time that data was downloaded. Each file can contain up to 50 tabs.
		ws_directory.cell(row=1, column=3).value = 'Date'	#Date this data was downloaded
		ws_directory.cell(row=1, column=4).value = 'Length'	#length (number of rows) fo the tab containing the data

	for n in range(len(Spinner_List)):
		new_directory = os.makedirs(Spinner_List[n])	#Creates a new folder for each spinner
		wb_data = Workbook()	#Creates a new excel workbook
		wb_data.save(Spinner_List[n] + '/' + Spinner_List[n] + '_1.xlsx')	#Saves this work book as "Spinner#_1.xlsx". This is the first.


	wb_directory.save('Wheel_Directory.xlsx')	#Save excel workbook

	print('\nFiles succesfully prepared\nReady to download running wheel data\n')	#prints that the file preparation is complete and that downloading can now occur.





def GETdata_wheel():

	for n in range(len(Spinner_List)):	#This loop will run through once for every spinner that you have.
		downloadWHEELdata(Spinner_List[n])	#Call function
		print('download process complete for ' + Spinner_List[n])	#Prints that the download is complete.

	os.chdir(Code_Directory)	#Directory where you can find your code

	length_directory()	#Calling function length_directory

	print('\ndownload process complete for all spinners from ' + Spinner_Email + '\n')	#Print message to know download was successful





def downloadWHEELdata(the_spinner):
	os.chdir(SavedData_Directory)	#Directory where the data will be saved.
	# Load the data and directory excel files

	Files = sorted(os.listdir(the_spinner))	#Create a sorted list of all the files found in this directory

	for n in range(len(Files)):
		if 'DS' in Files[n]:
			os.remove(the_spinner + '/' + Files[n])	# Delete the DS files so that they don't mess things ups.


	wb_data = load_workbook(the_spinner + '/' + Files[-1])	#Open the latest excel file from the spinner folder.
	w = 0
	wb_length = len(wb_data.sheetnames)	#This gives you the number of sheets in the file.

	#If the number of sheets is greater than or equal to 50, a new excel file will be created.
	#Need to do this or the MATLAB import will start to become extremely sluggish (openning and closing large excel files)
	if wb_length >= 50:
		wb_data = Workbook()
		w = w + 1

	wb_directory = load_workbook('Wheel_Directory.xlsx')	#Opens the directory excel file.
	
	ws_directory = wb_directory[the_spinner]	#Activate the sheet that corresponds with the appropriate wheel.

	#print('The subject is: ' + the_spinner)

	my_download = get_inbox(the_spinner)	#Calls the email function, passing it the subject to identify the appropriate wheel. This will return all the unread emails with that subject (spinner).

	email_patrol(my_download, the_spinner)	#Calls the email patrol function.

	# create my regular expressions
	nameRegex = re.compile(r'Sbinner\s\#\w\sReport')	#This will be found at the beginning of the email and tells you which spinner it is.
	dateRegex = re.compile(r'\w\w\w\w\-\w\w\-\w\w\s\w\w\:\w\w\:\w\w')	#This is the date and time at the start of this text file (when it was created).
	readRegex = re.compile(r'\'(\w{1,5}\.\w{1,3})\|(\w{1,3})\|(\w{1,3})\|(\w{1,3})\'')	#These are the individual entries which contain a time, and then values for left sensor, centre sensor, and right sensor.

	for n in range(len(my_download)):

		l = len(ws_directory['A'])	#Get the length of the directory, so that I can put the new values at the bottom of the directory.

		if "starting" not in my_download[n]['subject']:	#Prevents the starting/calibration emails from messing everything up.

			#Find all the corresponding values to the regular expressions that I created earlier.
			NAME = nameRegex.findall(my_download[n]['subject'])
			DATE = dateRegex.findall(my_download[n]['body'])
			READ = readRegex.findall(my_download[n]['body'])
		
			#Create and name the data sheet - naming it the date and time.
			ws_data = wb_data.create_sheet(DATE[0][0:10] + '_' + DATE[0][11:13] + DATE[0][14:16] + DATE[0][17:19], l-1)

			#Get the sheet name and the actual date (same things basically, but in different formats).
			ws_directory.cell(row=l+1, column=1).value = the_spinner + '_' + str(len(Files) + w) + '.xlsx'
			ws_directory.cell(row=l+1, column=2).value = DATE[0][0:10] + '_' + DATE[0][11:13] + DATE[0][14:16] + DATE[0][17:19]
			ws_directory.cell(row=l+1, column=3).value = DATE[0]
			#Create my Column names
			ws_data['A1'] = 'Time'
			ws_data['B1'] = 'Left_value'
			ws_data['C1'] = 'Centre_value'
			ws_data['D1'] = 'Right_value'
			#Use a loop to put all of my data into this table
			for x in range(len(READ)):
				m = x+2
				ws_data.cell(row=m, column=1).value = READ[x][0]
				ws_data.cell(row=m, column=2).value = READ[x][1]
				ws_data.cell(row=m, column=3).value = READ[x][2]
				ws_data.cell(row=m, column=4).value = READ[x][3]
			ws_directory.cell(row=l+1, column=4).value = len(ws_data['A'])

		#Checks to see how many sheets are in the workbook. If greater than or equal to 50 will create new workbook.
		if len(wb_data.sheetnames) >= 50:
			wb_data.save(the_spinner + '/' + the_spinner + '_' + str(len(Files) + w) + '.xlsx')

			wb_data = Workbook()
			w = w + 1


	#Saving my two excel files
	wb_data.save(the_spinner + '/' + the_spinner + '_' + str(len(Files) + w) + '.xlsx')	#The data file.
	wb_directory.save('Wheel_Directory.xlsx')	#The directory file.

	#Return my_download #For debugging purposes, otherwise keep blocked out.


def email_patrol(my_download, wheel):
	#This function checks that emails are still being sent from the spinner and sends an email alert if they are no longer being sent.

	#Check to see if there was an email received from that spinner.
	if len(my_download) < 1:	#If no email was receved from that spinner send an email alert.
		send_wheelalert(wheel)	#Calls function to send an email alert.
	
	#Checks time to see if an email was received in the past 4 hours (supposed to recieve an email every 3 hours)
	else:
		tz_UTC = pytz.timezone('UTC')
		tz_Alberta = pytz.timezone('US/Mountain')

		timeRegex = re.compile(r'\w\w\w\w\-\w\w\-\w\w\s\w\w\:\w\w\:\w\w')

		#The following for loop is necessary incase the last email sent was a calibration email. In which case there will be no time.
		Report_exists = False
		for n in range(len(my_download)):
			TIME = timeRegex.findall(my_download[-(n+1)][("subject")])
			if len(TIME) > 1:
				Report_exists = True
				Email_time = tz_UTC.localize(datetime.strptime(TIME[-1], "%Y-%m-%d %H:%M:%S"))
				NOW_time = datetime.now(tz_UTC)

				delta_time = NOW_time - Email_time
				delta_time_sec = delta_time.total_seconds()
				EmailAlert_seconds = EmailAlert_hours * 60 * 60	# 4 hours in seconds

				if delta_time_sec > EmailAlert_seconds:	#If an email was received more than 4 hours ago an alert email will be sent.
					send_wheelalert(wheel)
				break

		if Report_exists == False:
			send_wheelalert(wheel)





def length_directory():
	os.chdir(SavedData_Directory)	#Change folder directory to location of saved data.
	for n in range(len(Spinner_List)):	#Will run through this once for every spinner.
		wb_directory = load_workbook('Wheel_Directory.xlsx')	#Opens directory file.
		ws_length = wb_directory['Length']	#Opens sheet that has the lengths in it.
		ws_directory = wb_directory[Spinner_List[n]]	#Opens the directory sheet that corresponds with the spinner.

		Files = sorted(os.listdir(Spinner_List[n]))	#Creates a sorted list of all the files in that spinner's folder.

		ws_length.cell(row=n+2, column=2).value = len(Files)	#Records the number of files in the spinner's folder.
		ws_length.cell(row=n+2, column=3).value = len(ws_directory['A'])	#Records the number of sheets are in all files for that spinner.
		wb_directory.save('Wheel_Directory.xlsx')	#Saves the directory
