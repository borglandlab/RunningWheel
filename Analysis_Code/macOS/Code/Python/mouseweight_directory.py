import os
from openpyxl import load_workbook, Workbook

DIRECTORY = '/Users/<USERNAME>/Dropbox/RunningWheel'	#Location where the Mouse weight sheets are located
WORKBOOK = 'ABA.xlsx'

def mouseweight_directory():
	os.chdir(DIRECTORY)
	wb = load_workbook(WORKBOOK)

	sheet_list = wb.sheetnames
	sheet_list = sheet_list[0:-4]

	ws_directory = wb.create_sheet('Directory')
	ws_directory.cell(row=1, column=1).value = 'TabNumber'
	ws_directory.cell(row=1, column=2).value = 'TabName'
	ws_directory.cell(row=1, column=3).value = 'Date'
	ws_directory.cell(row=1, column=4).value = 'Experimental_Day'

	for n in range(len(sheet_list)):
		ws = wb[sheet_list[n]]
		ws_directory.cell(row=n+2, column=1).value = n+1
		ws_directory.cell(row=n+2, column=2).value = sheet_list[n]
		ws_directory.cell(row=n+2, column=3).value = ws.cell(row=3, column=2).value
		ws_directory.cell(row=n+2, column=4).value = ws.cell(row=2, column=2).value

	wb.save(WORKBOOK)