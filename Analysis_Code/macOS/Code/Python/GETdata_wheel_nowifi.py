import os, re, GETdata_wheel_nowifi
from openpyxl import load_workbook, Workbook
from datetime import datetime, timedelta

# Make sure this corresponds to the directory where I will find and save my data
SavedData_Directory = '/Users/<USERNAME>/Desktop/RunningWheel/Wheel_Data'	#This will need to be the folder directory where the data is saved.

#This will need to be as long as the number of spinners you have. This variable is used in Wheel_RunFirst and GetData_wheel functions.
Spinner_List = ['Spinner_1',
'Spinner_2',
'Spinner_3',
'Spinner_4',
'Spinner_5',
'Spinner_6',
'Spinner_7',
'Spinner_8']



def Wheel_RunFirst_nowifi():
	os.chdir(SavedData_Directory)

	wb_directory = Workbook()

	ws_length = wb_directory.create_sheet('Length', 0)
	ws_length.cell(row=1, column=1).value = 'Spinner'
	ws_length.cell(row=1, column=2).value = 'Files'
	ws_length.cell(row=1, column=3).value = 'Length'

	for n in range(len(Spinner_List)):
		ws_length.cell(row=n+2, column=1).value = Spinner_List[n]
		ws_directory = wb_directory.create_sheet(Spinner_List[n], n+1)
		ws_directory.cell(row=1, column=1).value = 'File'
		ws_directory.cell(row=1, column=2).value = 'Tab_Name'
		ws_directory.cell(row=1, column=3).value = 'Date'
		ws_directory.cell(row=1, column=4).value = 'Length'

	for n in range(len(Spinner_List)):
		wb_data = Workbook()
		wb_data.save(Spinner_List[n] + '/' + Spinner_List[n] + '_1.xlsx')


	wb_directory.save('Wheel_Directory.xlsx')

	print('\nFiles succesfully prepared\nReady to transfer running wheel data from text files\n')



def GETdata_wheel_nowifi():

	for n in range(len(Spinner_List)):
		TransferData_nowifi(Spinner_List[n])
		print('transfer process complete for ' + Spinner_List[n])

	os.chdir('/Users/nathansmac/MyPythonScripts')

	length_directory(Spinner_List)

	print('\ntransfer process complete for all spinners from borgland.mousebehaviour@gmail.com\n')	



def TransferData_nowifi(the_spinner):

	os.chdir(SavedData_Directory)

	# Load the data and directory excel files
	Files = sorted(os.listdir(the_spinner))

	for n in range(len(Files)):
		if 'DS' in Files[n]:
			os.remove(the_spinner + '/' + Files[n])

	wb_data = load_workbook(the_spinner + '/' + Files[-2])
	w = 0
	wb_length = len(wb_data.sheetnames)

	if wb_length >= 50:
		wb_data = Workbook()
		w = w + 1

	wb_directory = load_workbook('Wheel_Directory.xlsx')
	# Activate the sheet that corresponds with the appropriate wheel
	ws_directory = wb_directory[the_spinner]


	spinnerFile = open(the_spinner + "/spinlog_" + the_spinner[8:] + ".txt", "r")
	spinnerText = spinnerFile.read()

	# create my regular expressions
	nameRegex = re.compile(r'Spinner_\d{1,3}\sReport')

	dateRegex = re.compile(r'\w\w\w\w\-\w\w\-\w\w\s\w\w\:\w\w\:\w\w')

	readRegex = re.compile(r'\'(\w{1,5}\.\w{1,3})\|(\w{1,3})\|(\w{1,3})\|(\w{1,3})\'')

	DATE = dateRegex.findall(spinnerText)

	# Run through this for all but the last entry
	for n in range(len(DATE)-1):

		# get the length of the directory, so that I can put the new values at the bottom
		l = len(ws_directory['A'])	

		location_B = re.search(DATE[n], spinnerText)
		location_E = re.search(DATE[n+1], spinnerText)

		READ = readRegex.findall(spinnerText[location_B.span()[1]:location_E.span()[0]])

		ws_data = wb_data.create_sheet(DATE[n][0:10] + '_' + DATE[n][11:13] + DATE[n][14:16] + DATE[n][17:19], l-1)

		# get the sheet name and the actual date (same things basically, but in different formats)
		ws_directory.cell(row=l+1, column=1).value = the_spinner + '_' + str(len(Files) + w) + '.xlsx'
		ws_directory.cell(row=l+1, column=2).value = DATE[0][0:10] + '_' + DATE[0][11:13] + DATE[0][14:16] + DATE[0][17:19]
		ws_directory.cell(row=l+1, column=3).value = DATE[0]
		
		# create my variable name
		ws_data['A1'] = 'Time'
		ws_data['B1'] = 'Left_value'
		ws_data['C1'] = 'Centre_value'
		ws_data['D1'] = 'Right_value'		

		# use a loop to put all of my data into this table
		for x in range(len(READ)):
			m = x+2
			ws_data.cell(row=m, column=1).value = READ[x][0]
			ws_data.cell(row=m, column=2).value = READ[x][1]
			ws_data.cell(row=m, column=3).value = READ[x][2]
			ws_data.cell(row=m, column=4).value = READ[x][3]
		ws_directory.cell(row=l+1, column=4).value = len(ws_data['A'])

		if len(wb_data.sheetnames) >= 50:
			wb_data.save(the_spinner + '/' + the_spinner + '_' + str(len(Files) -1 + w) + '.xlsx')

			wb_data = Workbook()
			w = w + 1

	# Run through that again for the last entry
	l = len(ws_directory['A'])

	location_B = re.search(DATE[-1], spinnerText)

	READ = readRegex.findall(spinnerText[location_B.span()[1]:-1])

	ws_data = wb_data.create_sheet(DATE[n][0:10] + '_' + DATE[n][11:13] + DATE[n][14:16] + DATE[n][17:19], l-1)

	# get the sheet name and the actual date (same things basically, but in different formats)
	ws_directory.cell(row=l+1, column=1).value = the_spinner + '_' + str(len(Files) + w) + '.xlsx'
	ws_directory.cell(row=l+1, column=2).value = DATE[0][0:10] + '_' + DATE[0][11:13] + DATE[0][14:16] + DATE[0][17:19]
	ws_directory.cell(row=l+1, column=3).value = DATE[0]

	# create my variable name
	ws_data['A1'] = 'Time'
	ws_data['B1'] = 'Left_value'
	ws_data['C1'] = 'Centre_value'
	ws_data['D1'] = 'Right_value'	

	# use a loop to put all of my data into this table
	for x in range(len(READ)):
		m = x+2
		ws_data.cell(row=m, column=1).value = READ[x][0]
		ws_data.cell(row=m, column=2).value = READ[x][1]
		ws_data.cell(row=m, column=3).value = READ[x][2]
		ws_data.cell(row=m, column=4).value = READ[x][3]
	ws_directory.cell(row=l+1, column=4).value = len(ws_data['A'])

	if len(wb_data.sheetnames) >= 50:
		wb_data.save(the_spinner + '/' + the_spinner + '_' + str(len(Files) -1 + w) + '.xlsx')

		wb_data = Workbook()
		w = w + 1

	# Saving my two excel files
	wb_data.save(the_spinner + '/' + the_spinner + '_' + str(len(Files) - 1 + w) + '.xlsx')
	wb_directory.save('Wheel_Directory.xlsx')



def length_directory(the_spinner):
	os.chdir(SavedData_Directory)
	for n in range(len(the_spinner)):
		wb_directory = load_workbook('Wheel_Directory.xlsx')
		ws_length = wb_directory['Length']
		ws_directory = wb_directory[the_spinner[n]]

		Files = sorted(os.listdir(the_spinner[n]))

		ws_length.cell(row=n+2, column=2).value = len(Files)
		ws_length.cell(row=n+2, column=3).value = len(ws_directory['A'])
		wb_directory.save('Wheel_Directory.xlsx')


